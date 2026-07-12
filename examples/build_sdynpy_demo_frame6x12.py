# -*- coding: utf-8 -*-
"""
build_sdynpy_demo_frame6x12.py

Builds a larger synthetic MIMO test article than build_sdynpy_demo.py's
5-mass chain: a 2D lumped-mass frame with 6 shaker drive points and 12
accelerometer response points, tuned so essentially all of its flexible
modes fall inside the 100-1000 Hz band. Exports a native SDynPy `System`
plus a fully populated Rattlesnake profile spreadsheet, the same way
build_sdynpy_demo.py does for the small demo.

System layout: 3 rows x 6 columns = 18 nodes, each with 2 in-plane DOFs
(X, Y), connected by row springs (axial, within each row), riser springs
(vertical, between rows), and diagonal braces (between rows, offset one
column) so the structure has genuine 2D/cross-axis coupling rather than
being a simple 1D chain:

    Row A (y=0): [1]--[2]--[3]--[4]--[5]--[6]      <- 6 shaker drives (X)
                  |  X  |  X  |  X  |  X  |
    Row B (y=1): [7]--[8]--[9]--[10]-[11]-[12]     <- 6 accelerometers (X)
                  |  X  |  X  |  X  |  X  |
    Row C (y=2): [13]-[14]-[15]-[16]-[17]-[18]     <- 6 accelerometers (X)

  * 6 shaker inputs: force (X) at every Row A node
  * 12 accelerometer outputs: acceleration (X) at every Row B/C node

Stiffness is scaled so the lowest flexible mode sits at ~120 Hz and the
highest at ~900 Hz -- all 33 flexible modes land in the 100-1000 Hz band,
giving Rattlesnake's MIMO Random control law (and any control law being
compared against it) plenty of real resonant, cross-coupled content to
work with.

The generated profile pre-fills the "Random" environment tab with a
complete, ready-to-run configuration:
  * Control law = control_laws/control_laws.py :: buzz_control_class
    (the standard pseudo-inverse MIMO Random law). To compare a different
    control law, either edit the "Control Python Script" / "Control
    Python Function" cells on the environment sheet before loading, or
    change them in the GUI after loading -- e.g. point at
    control_laws/optimal_diagonal_control.py to compare against the SDP-
    based optimal diagonal law.
  * Control channels = the first 8 of the 12 response channels (all 6 of
    Row B + first 2 of Row C). The other 4 response channels are still
    in the channel table and active in the environment (so they show up
    as extra response/monitor channels) -- check more of them in the
    "Control Channels" list in the environment definition GUI at any
    time to expand from 8 up toward all 12, no need to rebuild the model.
  * Specification file = flat_spec_frame6x12.mat (see
    build_flat_spec_large.py), a flat 100-1000 Hz CPSD target sized for
    the 8 default control channels. Regenerate that file with a
    different Nc if you change the control channel count.

Run (from the `sdynpy` conda environment -- needs sdynpy, qtpy, pyqtgraph,
openpyxl):

    conda activate sdynpy
    cd ~/Documents/Code/python/rattlesnake-vibration-controller/examples
    python build_sdynpy_demo_frame6x12.py
    python build_flat_spec_large.py

Then in Rattlesnake: File -> Open Profile -> sdynpy_frame6x12_profile.xlsx.
"""

import os
import numpy as np
from scipy.linalg import eigh
import openpyxl as opxl
import sdynpy as sdpy
from sdynpy.fileio.sdynpy_rattlesnake import create_synthetic_test

# ---------------------------------------------------------------------
# 1. Grid topology: 3 rows x 6 columns = 18 nodes, 2 DOF/node (X, Y)
# ---------------------------------------------------------------------
n_cols = 6
n_rows = 3  # row 0 = drive row (A), rows 1-2 = response rows (B, C)
n_nodes = n_cols * n_rows
ndof = n_nodes * 2  # (x, y) per node


def node_number(row, col):
    """1-based node number, row-major (matches the docstring diagram)."""
    return row * n_cols + col + 1


def node_index(row, col):
    """0-based node index into the mass/stiffness matrices."""
    return row * n_cols + col


def dofs(row, col):
    i = node_index(row, col)
    return 2 * i, 2 * i + 1


# Node positions (row spacing 1.0, column spacing 1.0) -- only used to get
# spring direction cosines, not absolute physical dimensions.
pos = np.zeros((n_nodes, 2))
for r in range(n_rows):
    for c in range(n_cols):
        pos[node_index(r, c)] = [c * 1.0, r * 1.0]

# ---------------------------------------------------------------------
# 2. Stiffness matrix: row springs (axial), riser springs (vertical),
#    and diagonal braces (2D coupling), each with mild deterministic
#    non-uniformity so modes aren't degenerate/symmetric.
# ---------------------------------------------------------------------
K_unit = np.zeros((ndof, ndof))
_spring_counter = [0]


def kvar(base):
    """Deterministic +/-6% variation, different for every spring."""
    i = _spring_counter[0]
    _spring_counter[0] += 1
    return base * (1.0 + 0.06 * np.sin(1.7 * i + 0.3))


def add_spring(node_a, node_b, k):
    ia, ja = dofs(*node_a)
    ib, jb = dofs(*node_b)
    d = pos[node_index(*node_b)] - pos[node_index(*node_a)]
    length = np.linalg.norm(d)
    c, s = d / length
    t = np.array([c, s, -c, -s])
    local = k * np.outer(t, t)
    idx = [ia, ja, ib, jb]
    for a in range(4):
        for b in range(4):
            K_unit[idx[a], idx[b]] += local[a, b]


k_row_drive = 1.0  # Row A (drive row) axial stiffness
k_row = 1.0        # Row B, C (response rows) axial stiffness
k_riser = 1.0       # vertical stiffness between rows
k_diag = 0.6        # diagonal brace stiffness (2D coupling)

# Row springs (axial, within each row)
for r in range(n_rows):
    base = k_row_drive if r == 0 else k_row
    for c in range(n_cols - 1):
        add_spring((r, c), (r, c + 1), kvar(base))

# Riser springs (vertical, between adjacent rows, same column)
for c in range(n_cols):
    add_spring((0, c), (1, c), kvar(k_riser))
    add_spring((1, c), (2, c), kvar(k_riser))

# Diagonal braces (between adjacent rows, offset one column -- the source
# of genuine 2D/cross-axis coupling beyond a simple ladder)
for c in range(n_cols - 1):
    add_spring((0, c), (1, c + 1), kvar(k_diag))
    add_spring((1, c), (2, c + 1), kvar(k_diag))

# ---------------------------------------------------------------------
# 3. Mass matrix: mild deterministic non-uniformity
# ---------------------------------------------------------------------
m = np.array([0.5 + 0.04 * np.sin(0.9 * i + 0.5) for i in range(n_nodes)])
M = np.zeros((ndof, ndof))
for i in range(n_nodes):
    M[2 * i, 2 * i] = m[i]
    M[2 * i + 1, 2 * i + 1] = m[i]

# ---------------------------------------------------------------------
# 4. Scale K_unit so the lowest flexible mode lands at target_low Hz
#    (free-free structure: first 3 modes are rigid body, ~0 Hz)
# ---------------------------------------------------------------------
n_rigid = 3
target_low_hz = 120.0

eigvals_unit, _ = eigh(K_unit, M)
eigvals_unit = np.clip(eigvals_unit, 0, None)
freqs_unit = np.sqrt(eigvals_unit) / (2 * np.pi)
flex_unit = freqs_unit[n_rigid:]

scale = (target_low_hz / flex_unit.min()) ** 2
K = K_unit * scale

# ---------------------------------------------------------------------
# 5. Rayleigh damping: C = alpha*M + beta*K, tuned to ~zeta_target at
#    two representative flexible modes spanning the band
# ---------------------------------------------------------------------
zeta_target = 0.01

eigvals, _ = eigh(K, M)
eigvals = np.clip(eigvals, 0, None)
omega_n = np.sqrt(eigvals)
freqs_hz = omega_n / (2 * np.pi)

i1 = n_rigid
i2 = n_rigid + (len(freqs_hz) - n_rigid) * 2 // 3
w1, w2 = omega_n[i1], omega_n[i2]
A_ray = 0.5 * np.array([[1 / w1, w1], [1 / w2, w2]])
alpha, beta = np.linalg.solve(A_ray, [zeta_target, zeta_target])
C_damp = alpha * M + beta * K

# ---------------------------------------------------------------------
# 6. Coordinates: full 36-DOF system coordinate array (X, Y per node),
#    then the excitation (Row A, X only) and response (Rows B/C, X
#    only) subsets that actually get exposed to Rattlesnake.
# ---------------------------------------------------------------------
all_node_numbers = [node_number(r, c) for r in range(n_rows) for c in range(n_cols)]
full_coordinate = sdpy.coordinate_array(
    node=np.repeat(all_node_numbers, 2),
    direction=np.tile(['X', 'Y'], n_nodes),
)

drive_node_numbers = [node_number(0, c) for c in range(n_cols)]
response_node_numbers = (
    [node_number(1, c) for c in range(n_cols)]  # Row B
    + [node_number(2, c) for c in range(n_cols)]  # Row C
)

excitation_coordinates = sdpy.coordinate_array(node=drive_node_numbers, direction='X')
response_coordinates = sdpy.coordinate_array(node=response_node_numbers, direction='X')

n_control_channels = 8  # first 8 of the 12 response channels; adjustable in the GUI later

# ---------------------------------------------------------------------
# 7. Build the SDynPy System and generate the system file + profile
# ---------------------------------------------------------------------
system = sdpy.System(full_coordinate, mass=M, stiffness=K, damping=C_damp)

output_dir = os.path.dirname(os.path.abspath(__file__))
system_filename = os.path.join(output_dir, "sdynpy_frame6x12_system.npz")
spreadsheet_file_name = os.path.join(output_dir, "sdynpy_frame6x12_profile.xlsx")
spec_filename = os.path.join(output_dir, "flat_spec_frame6x12.mat")

rattlesnake_directory = os.path.expanduser(
    "~/Documents/Code/python/rattlesnake-vibration-controller"
)
control_law_script = os.path.join(rattlesnake_directory, "control_laws", "control_laws.py")

sample_rate = 5120          # Hz -- ~5.7x the ~900 Hz top flexible mode
samples_per_frame = 5120    # 1 s frames -> 1 Hz frequency resolution
time_per_read = 1.0
time_per_write = 1.0

environment_name = "Frame 6x12 Random"

create_synthetic_test(
    spreadsheet_file_name=spreadsheet_file_name,
    system_filename=system_filename,
    system=system,
    excitation_coordinates=excitation_coordinates,
    response_coordinates=response_coordinates,
    rattlesnake_directory=rattlesnake_directory,
    sample_rate=sample_rate,
    time_per_read=time_per_read,
    time_per_write=time_per_write,
    environments=[("random", environment_name)],
)

# ---------------------------------------------------------------------
# 8. Fill in the "Random" environment sheet with a real, ready-to-run
#    configuration (create_synthetic_test only populates the Channel
#    Table and Hardware sheets, leaving the environment sheet as a
#    blank template).
# ---------------------------------------------------------------------
workbook = opxl.load_workbook(spreadsheet_file_name)
env_ws = workbook[environment_name]

env_ws.cell(2, 2, samples_per_frame)
env_ws.cell(3, 2, 2.0)          # Test Level Ramp Time (s)
env_ws.cell(4, 2, "Hann")       # COLA Window
env_ws.cell(5, 2, 50.0)         # COLA Overlap %
env_ws.cell(6, 2, 0.5)          # COLA Window Exponent
env_ws.cell(7, 2, "N")          # Update System ID During Control
env_ws.cell(8, 2, 10)           # Frames in CPSD
env_ws.cell(9, 2, "Hann")       # CPSD Window
env_ws.cell(10, 2, 50.0)        # CPSD Overlap %
env_ws.cell(11, 2, "N")         # Allow Automatic Aborts
env_ws.cell(12, 2, control_law_script)     # Control Python Script
env_ws.cell(13, 2, "buzz_control_class")   # Control Python Function
env_ws.cell(14, 2, "")          # Control Parameters
for col_offset, channel_index in enumerate(range(1, n_control_channels + 1)):
    env_ws.cell(15, 2 + col_offset, channel_index)  # Control Channels (1-based)
env_ws.cell(16, 2, "Linear")    # System ID Averaging
env_ws.cell(17, 2, 0)           # Noise Averages
env_ws.cell(18, 2, 5)           # System ID Averages
env_ws.cell(19, 2, 0.1)         # Exponential Averaging Coefficient
env_ws.cell(20, 2, "H1")        # System ID Estimator
env_ws.cell(21, 2, 1.0)         # System ID Level (V RMS)
env_ws.cell(22, 2, "Random")    # System ID Signal Type
env_ws.cell(23, 2, "Hann")      # System ID Window
env_ws.cell(24, 2, 50.0)        # System ID Overlap %
env_ws.cell(25, 2, 100.0)       # System ID Burst On %
env_ws.cell(26, 2, 0.0)         # System ID Burst Pretrigger %
env_ws.cell(27, 2, 0.0)         # System ID Ramp Fraction %
env_ws.cell(28, 2, spec_filename)  # Specification File
env_ws.cell(29, 2, "None")      # Response Transformation Matrix
env_ws.cell(30, 2, "None")      # Output Transformation Matrix

workbook.save(spreadsheet_file_name)

# ---------------------------------------------------------------------
# 9. Sanity check printout
# ---------------------------------------------------------------------
print(f"System file written to:      {system_filename}")
print(f"Profile spreadsheet written: {spreadsheet_file_name}")
print()
print(f"{len(drive_node_numbers)} shaker inputs at nodes {drive_node_numbers}")
print(f"{len(response_node_numbers)} accelerometer outputs at nodes {response_node_numbers}")
print(f"{n_control_channels} of those marked as control channels (indices 1-{n_control_channels})")
print()
print("Natural frequencies (Hz):")
flex = freqs_hz[n_rigid:]
for i, f in enumerate(freqs_hz):
    tag = "(rigid body)" if i < n_rigid else ""
    print(f"  Mode {i + 1:2d}: {f:8.2f} Hz  {tag}")
in_band = (flex >= 100) & (flex <= 1000)
print()
print(f"{in_band.sum()} / {len(flex)} flexible modes land in 100-1000 Hz "
      f"(range: {flex.min():.1f} - {flex.max():.1f} Hz)")
print()
print(f"Rayleigh damping: alpha={alpha:.6e}, beta={beta:.6e}")
print(f"Target zeta={zeta_target:.3f} enforced at modes {i1 + 1} and {i2 + 1}")
print()
print("Next: python build_flat_spec_large.py, then in Rattlesnake:")
print("File -> Open Profile ->", spreadsheet_file_name)
